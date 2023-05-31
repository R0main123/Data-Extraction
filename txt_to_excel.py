import timeit
import os
import pandas as pd
from openpyxl.reader.excel import load_workbook


def spliter(line):
    """
    This function takes a line of a header from a .txt file in argument and returns the relevant information in this line
    :param <str> line: The line you want to extract the information
    :return: the information needed
    :rtype: str
    """
    return line.split(' : ')[-1][:-1]

def dataSpliter(line):
    """
    This function takes a line of datas from a .txt file in argument and returns a list with voltage in first position and current in second position
    :param <str> line: The line you want to extract the information
    :return: the information needed
    :rtype: list of str
    """
    return line[:-1].split('\t')

def convert_to_excel(path):
    """
    This function takes a path to a .txt file in argument and creates an excel files with the values of current depending on corresponding voltage
    :param <str> path: The path to the .txt file
    :return: None
    """
    start_time = timeit.default_timer()

    #creating directory if it doesn't exist
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    #Creating Excel Files: 1 for positives values, 1 for negatives values
    filename = 'ExcelFiles\\'+path.split('/')[-1].split('.')[0]+'_Data_IV.xlsx'
    df = pd.DataFrame()
    df.to_excel(filename, index = False)

    #Creating lists of sheets that already exists in the excel files
    list_of_sheets = []

    #Processing files
    with open(path, 'r') as file:
        i=1
        while True:
            start_iter = timeit.default_timer()

            #Initilization of columns
            df = pd.DataFrame()
            voltages = ['Voltage (V)']
            I = ['I (A)']

            #collection of relevant data
            line = next((l for l in file if 'chipX' in l), None)
            if not line:
                break
            chipX = spliter(line)

            line = next((l for l in file if 'chipY' in l), None)
            if not line:
                break
            chipY = spliter(line)

            line = next((l for l in file if 'testdeviceID' in l), None)
            if not line:
                break
            testdeviceID = spliter(line)

            #Creating new sheet
            sheetName = "(" + str(chipX) + "," + str(chipY) + ")"

            line = next((l for l in file if 'BOD' in l), None)
            if not line:
                break

            #Collecting measures values
            for line in file:
                if 'EOD' in line:
                    break
                data = dataSpliter(line)
                voltages.append(data[0])
                I.append(data[-1])


            #Getting the relevent column
            df = pd.DataFrame(list(zip(voltages, I)), columns=[testdeviceID, testdeviceID])


            #Writing datas in the right file

            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if sheetName in list_of_sheets:
                    df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
                else:
                    df.to_excel(writer, sheet_name=sheetName, index=False)
                    list_of_sheets.append(sheetName)

            end_iter = timeit.default_timer()
            print(f"Iteration number {i} ended in {end_iter - start_iter} seconds.")
            i+=1

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    wb.save(filename)

    end_time = timeit.default_timer()
    execution_time = end_time - start_time

    print(f"Ended in {execution_time} secondes")