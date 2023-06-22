import timeit
import os
import pandas as pd
from openpyxl.reader.excel import load_workbook
from pymongo import MongoClient

def Excel_IV(structure_id=str, values=list,wafer=str, df_dict=dict, filename=str, list_of_sheets=list):
    """
    This function takes a path to a .txt file in argument and creates an excel files with the values of current depending on corresponding voltage
    :param <str> path: The path to the .txt file
    :return: None
    """

    for structure in wafer["structures"]:
        for matrix in structure["matrices"]:
            coord = "("+matrix["coordinates"]["x"]+','+matrix["coordinates"]["y"]+')'

            if coord not in df_dict.keys():
                df_dict[coord] = pd.DataFrame()

            testdeviceID = structure["structure_id"]
            voltages = ['Voltage (V)']
            I = ['I (A)']
            for double in matrix["results"]["I"]["Values"]:
                voltages.append(double["V"])
                I.append(double["I"])

            new_df = pd.DataFrame(list(zip(voltages, I)), columns=[testdeviceID, testdeviceID+" "])
            df_dict[coord] = df_dict[coord].merge(new_df, left_index=True, right_index=True, how = 'outer')

            #i += 1

    pd.DataFrame().to_excel(filename, index=False)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for coord, df in df_dict.items():
            sheetName = coord
            if sheetName in list_of_sheets:
                df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
            else:
                df.to_excel(writer, sheet_name=sheetName, index=False)
                list_of_sheets.append(sheetName)

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    wb.save(filename)

    end_time = timeit.default_timer()


    #print(f"I-V Excel successfully Ended in {execution_time} seconds for {wafer_id} ({i} iterations)")

def Excel_JV(wafer_id=str):
    """
    This function takes a path to a .txt file in argument and creates an excel files with the values of current depending on corresponding voltage
    :param <str> path: The path to the .txt file
    :return: None
    """
    if os.path.exists(f"ExcelFiles\\{wafer_id}_JV.xlsx"):
        return
    start_time = timeit.default_timer()

    client = MongoClient('mongodb://localhost:27017/')
    db = client['Measurements']
    collection = db["Wafers"]
    wafer = collection.find_one({"wafer_id": wafer_id})

    # creating directory if it doesn't exist
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    # Creating Excel File
    filename = 'ExcelFiles\\' + wafer_id + '_JV.xlsx'

    # Creating lists of sheets that already exists in the excel files
    list_of_sheets = []
    df_dict = {}  # using dictionary to store dataframes by coord keys

    # Processing files
    i = 1
    for structure in wafer["structures"]:
        for matrix in structure["matrices"]:
            coord = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'

            if coord not in df_dict.keys():
                df_dict[coord] = pd.DataFrame()

            testdeviceID = structure["structure_id"]
            voltages = ['Voltage (V)']
            I = ['J (A/cm^2)']
            for double in matrix["results"]["J"]["Values"]:
                voltages.append(double["V"])
                I.append(double["J"])

            new_df = pd.DataFrame(list(zip(voltages, I)), columns=[testdeviceID, testdeviceID + " "])
            df_dict[coord] = df_dict[coord].merge(new_df, left_index=True, right_index=True, how='outer')

            end_iter = timeit.default_timer()
            i += 1

    pd.DataFrame().to_excel(filename, index=False)

    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for coord, df in df_dict.items():
            sheetName = coord
            if sheetName in list_of_sheets:
                df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
            else:
                df.to_excel(writer, sheet_name=sheetName, index=False)
                list_of_sheets.append(sheetName)

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    wb.save(filename)

    end_time = timeit.default_timer()
    execution_time = end_time - start_time

    print(f"J-V excel successfully created for {wafer_id}. Ended in {execution_time} secondes")

def Excel_CV(wafer_id=str):
    """
     This function takes a path to a .txt file in argument and creates an excel files with the values of current depending on corresponding voltage
     :param <str> path: The path to the .txt file
     :return: None
     """

    if os.path.exists(f"ExcelFiles\\{wafer_id}_CV.xlsx"):
        return
    start_time = timeit.default_timer()

    client = MongoClient('mongodb://localhost:27017/')
    db = client['Measurements']
    collection = db["Wafers"]
    wafer = collection.find_one({"wafer_id": wafer_id})

    # creating directory if it doesn't exist
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    # Creating Excel File
    filename = 'ExcelFiles\\' + wafer_id + '_CV.xlsx'

    # Creating lists of sheets that already exists in the excel files
    list_of_sheets = []
    df_dict = {}  # using dictionary to store dataframes by coord keys

    # Processing files
    i = 1
    for structure in wafer["structures"]:
        for matrix in structure["matrices"]:
            coord = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'

            if coord not in df_dict.keys():
                df_dict[coord] = pd.DataFrame()

            testdeviceID = structure["structure_id"]
            voltages = ['Voltage (V)']
            CS = ['CS (Ω)']
            RS = ['RS (Ω)']
            for double in matrix["results"]["C"]["Values"]:
                voltages.append(double["V"])
                CS.append(double["CS"])
                RS.append(double["RS"])

            new_df = pd.DataFrame(list(zip(voltages, CS, RS)), columns=[testdeviceID, testdeviceID + " ", testdeviceID+ "  "])
            df_dict[coord] = df_dict[coord].merge(new_df, left_index=True, right_index=True, how='outer')

            i += 1

    pd.DataFrame().to_excel(filename, index=False)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for coord, df in df_dict.items():
            sheetName = coord
            if sheetName in list_of_sheets:
                df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
            else:
                df.to_excel(writer, sheet_name=sheetName, index=False)
                list_of_sheets.append(sheetName)

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    wb.save(filename)

    end_time = timeit.default_timer()
    execution_time = end_time - start_time

    print(f"C-V Excel successfully Ended in {execution_time} seconds for {wafer_id} ({i} iterations)")

def Excel_TDDB(wafer_id=str):
    """
    This function takes a path to a .txt file in argument and creates an excel files with the values of current depending on corresponding voltage
    :param <str> path: The path to the .txt file
    :return: None
    """

    if os.path.exists(f"ExcelFiles\\{wafer_id}_TDDB.xlsx"):
        return
    start_time = timeit.default_timer()

    client = MongoClient('mongodb://localhost:27017/')
    db = client['Measurements']
    collection = db["Wafers"]
    wafer = collection.find_one({"wafer_id": wafer_id})

    #creating directory if it doesn't exist
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    #Creating Excel File
    filename = 'ExcelFiles\\'+wafer_id+'_TDDB.xlsx'

    #Creating lists of sheets that already exists in the excel files
    list_of_sheets = []
    df_dict = {}  # using dictionary to store dataframes by coord keys

    #Processing files
    i = 1
    for structure in wafer["structures"]:
        for matrix in structure["matrices"]:
            coord = "("+matrix["coordinates"]["x"]+','+matrix["coordinates"]["y"]+')'

            if coord not in df_dict.keys():
                df_dict[coord] = pd.DataFrame()

            testdeviceID = structure["structure_id"]
            voltages = ['Voltage (V)']
            It = ['TDDB']
            for double in matrix["results"]["It"]["Values"]:
                voltages.append(double["V"])
                It.append(double["It"])

            new_df = pd.DataFrame(list(zip(voltages, It)), columns=[testdeviceID, testdeviceID+" "])
            df_dict[coord] = df_dict[coord].merge(new_df, left_index=True, right_index=True, how = 'outer')

            i += 1

    pd.DataFrame().to_excel(filename, index=False)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for coord, df in df_dict.items():
            sheetName = coord
            if sheetName in list_of_sheets:
                df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
            else:
                df.to_excel(writer, sheet_name=sheetName, index=False)
                list_of_sheets.append(sheetName)

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    wb.save(filename)

    end_time = timeit.default_timer()
    execution_time = end_time - start_time

    print(f"TDDB Excel successfully Ended in {execution_time} seconds for {wafer_id} ({i} iterations)")

def writeExcel(wafer):
    print("Starting excel...")
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    client = MongoClient('mongodb://localhost:27017/')
    db = client['Measurements']
    collection = db["Wafers"]

    wafer = collection.find_one({"wafer_id": wafer})
    list_of_sheets = []
    df_dict = {}


    wafer_id = wafer["wafer_id"]
    df_dict["I"] = {}
    df_dict["J"] = {}
    df_dict["C"] = {}
    df_dict["It"] = {}


    for structure in wafer["structures"]:
        structure_id = structure["structure_id"]

        for matrix in structure["matrices"]:
            coord = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'
            if coord not in df_dict.keys():
                df_dict[coord] = pd.DataFrame()

            for element in matrix["results"]:
                if element == "I":
                    if os.path.exists(f"ExcelFiles\\{wafer_id}_IV.xlsx"):
                        continue
                    if coord not in df_dict["I"]:
                        df_dict["I"][coord] = pd.DataFrame()

                    voltages = ['Voltage (V)']
                    I = ['I (A)']
                    for double in matrix["results"][element]["Values"]:
                        voltages.append(double["V"])
                        I.append(double["I"])

                    new_df = pd.DataFrame(list(zip(voltages, I)), columns=[structure_id, structure_id + " "])
                    df_dict["I"][coord] = df_dict["I"][coord].merge(new_df, left_index=True, right_index=True, how='outer')
                    I.clear()
                    voltages.clear()

                elif element == "J":
                    if coord not in df_dict["J"]:
                        df_dict["J"][coord] = pd.DataFrame()
                    if os.path.exists(f"ExcelFiles\\{wafer_id}_JV.xlsx"):
                        continue
                    voltages = ['Voltage (V)']
                    J = ['J (A/cm^2)']
                    for double in matrix["results"][element]["Values"]:
                        voltages.append(double["V"])
                        J.append(double["J"])

                    new_df = pd.DataFrame(list(zip(voltages, J)), columns=[structure_id, structure_id + " "])
                    df_dict["J"][coord] = df_dict["J"][coord].merge(new_df, left_index=True, right_index=True,
                                                          how='outer')
                    J.clear()
                    voltages.clear()

                elif element == "C":
                    if os.path.exists(f"ExcelFiles\\{wafer_id}_CV.xlsx"):
                        continue

                    if coord not in df_dict["C"]:
                        df_dict["C"][coord] = pd.DataFrame()
                    voltages = ['Voltage (V)']
                    CS = ['CS (Ω)']
                    RS = ['RS (Ω)']
                    for double in matrix["results"][element]["Values"]:
                        voltages.append(double["V"])
                        CS.append(double["CS"])
                        RS.append(double["RS"])

                    new_df = pd.DataFrame(list(zip(voltages, CS, RS)),
                                          columns=[structure_id, structure_id + " ", structure_id + "  "])
                    df_dict["C"][coord] = df_dict["C"][coord].merge(new_df, left_index=True, right_index=True,
                                                          how='outer')
                    voltages.clear()
                    CS.clear()
                    RS.clear()

                elif element == "It":
                    if os.path.exists(f"ExcelFiles\\{wafer_id}_TDDB.xlsx"):
                        continue
                    if coord not in df_dict["It"]:
                        df_dict["It"][coord] = pd.DataFrame()
                    voltages = ['Voltage (V)']
                    It = ['TDDB']
                    for double in matrix["results"][element]["Values"]:
                        voltages.append(double["V"])
                        It.append(double["TDDB"])

                    new_df = pd.DataFrame(list(zip(voltages, It)), columns=[structure_id, structure_id + " "])
                    df_dict["It"][coord] = df_dict["It"][coord].merge(new_df, left_index=True, right_index=True,
                                                          how='outer')
                    voltages.clear()
                    It.clear()

    for element in ["I","J","C","It"]:
        if df_dict[element] != {}:
            filename = 'ExcelFiles\\' + wafer_id + "_" +element +'V.xlsx'

            pd.DataFrame().to_excel(filename, index=False)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for coord, df in df_dict[element].items():
                    sheetName = coord
                    if sheetName in list_of_sheets:
                        df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
                    else:
                        df.to_excel(writer, sheet_name=sheetName, index=False)
                        list_of_sheets.append(sheetName)
            df_dict[element].clear()
            list_of_sheets.clear()

            wb = load_workbook(filename)
            if 'Sheet1' in wb.sheetnames:
                del wb['Sheet1']
            wb.save(filename)

def excel_structure(wafer_id=str, structure_ids=list, file_name=str):
    if not os.path.exists("ExcelFiles"):
        os.makedirs("ExcelFiles")

    client = MongoClient('mongodb://localhost:27017/')
    db = client['Measurements']
    collection = db["Wafers"]

    wafer = collection.find_one({"wafer_id": wafer_id})
    list_of_sheets = []
    df_dict = {}

    df_dict["I"] = {}
    df_dict["J"] = {}
    df_dict["C"] = {}
    df_dict["It"] = {}

    for structure in wafer["structures"]:
        if structure["structure_id"] in structure_ids:
            structure_id = structure["structure_id"]

            for matrix in structure["matrices"]:
                coord = "(" + matrix["coordinates"]["x"] + ',' + matrix["coordinates"]["y"] + ')'
                if coord not in df_dict.keys():
                    df_dict[coord] = pd.DataFrame()

                for element in matrix["results"]:
                    if element == "I":
                        if os.path.exists(f"ExcelFiles\\{file_name}_I-V.xlsx"):
                            continue
                        if coord not in df_dict["I"]:
                            df_dict["I"][coord] = pd.DataFrame()

                        voltages = ['Voltage (V)']
                        I = ['I (A)']
                        for double in matrix["results"][element]["Values"]:
                            voltages.append(double["V"])
                            I.append(double["I"])

                        new_df = pd.DataFrame(list(zip(voltages, I)), columns=[structure_id, structure_id + " "])
                        df_dict["I"][coord] = df_dict["I"][coord].merge(new_df, left_index=True, right_index=True, how='outer')
                        I.clear()
                        voltages.clear()

                    elif element == "J":
                        if coord not in df_dict["J"]:
                            df_dict["J"][coord] = pd.DataFrame()
                        if os.path.exists(f"ExcelFiles\\{file_name}_J-V.xlsx"):
                            continue
                        voltages = ['Voltage (V)']
                        J = ['J (A/cm^2)']
                        for double in matrix["results"][element]["Values"]:
                            voltages.append(double["V"])
                            J.append(double["J"])

                        new_df = pd.DataFrame(list(zip(voltages, J)), columns=[structure_id, structure_id + " "])
                        df_dict["J"][coord] = df_dict["J"][coord].merge(new_df, left_index=True, right_index=True,
                                                              how='outer')
                        J.clear()
                        voltages.clear()

                    elif element == "C":
                        if os.path.exists(f"ExcelFiles\\{file_name}_C-V.xlsx"):
                            continue

                        if coord not in df_dict["C"]:
                            df_dict["C"][coord] = pd.DataFrame()
                        voltages = ['Voltage (V)']
                        CS = ['CS (Ω)']
                        RS = ['RS (Ω)']
                        for double in matrix["results"][element]["Values"]:
                            voltages.append(double["V"])
                            CS.append(double["CS"])
                            RS.append(double["RS"])

                        new_df = pd.DataFrame(list(zip(voltages, CS, RS)),
                                              columns=[structure_id, structure_id + " ", structure_id + "  "])
                        df_dict["C"][coord] = df_dict["C"][coord].merge(new_df, left_index=True, right_index=True,
                                                              how='outer')
                        voltages.clear()
                        CS.clear()
                        RS.clear()

                    elif element == "It":
                        if os.path.exists(f"ExcelFiles\\{file_name}_It-V.xlsx"):
                            continue
                        if coord not in df_dict["It"]:
                            df_dict["It"][coord] = pd.DataFrame()
                        voltages = ['Voltage (V)']
                        It = ['TDDB']
                        for double in matrix["results"][element]["Values"]:
                            voltages.append(double["V"])
                            It.append(double["TDDB"])

                        new_df = pd.DataFrame(list(zip(voltages, It)), columns=[structure_id, structure_id + " "])
                        df_dict["It"][coord] = df_dict["It"][coord].merge(new_df, left_index=True, right_index=True,
                                                              how='outer')
                        voltages.clear()
                        It.clear()

    for element in ["I","J","C","It"]:
        if df_dict[element] != {}:

            filename = 'ExcelFiles\\' + file_name + "_" +element +'-V.xlsx'

            pd.DataFrame().to_excel(filename, index=False)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for coord, df in df_dict[element].items():
                    sheetName = coord
                    if sheetName in list_of_sheets:
                        df.to_excel(writer, sheet_name=sheetName, index=False, startcol=writer.sheets[sheetName].max_column)
                    else:
                        df.to_excel(writer, sheet_name=sheetName, index=False)
                        list_of_sheets.append(sheetName)
            df_dict[element].clear()
            list_of_sheets.clear()

            wb = load_workbook(filename)
            if 'Sheet1' in wb.sheetnames:
                del wb['Sheet1']
            wb.save(filename)
